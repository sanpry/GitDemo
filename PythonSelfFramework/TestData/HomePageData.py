import openpyxl
from Tools.scripts.fixcid import Dict


class HomePageData:

    test_HomePage_data = [
        {"name": "swe", "email": "ssp@gmail.com", "pwd": "123", "gender": "Female", "status": "Employed",
         "dob": "16-12-1990"},
        {"name": "sangeetha", "email": "ssp@gmail.com", "pwd": "123", "gender": "Female", "status": "Employed",
         "dob": "16-12-1990"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\swetha.priyanka.IDSNEXT\\Desktop\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                    # print(sheet.cell(row=i, column=j).value)
        return[Dict]
