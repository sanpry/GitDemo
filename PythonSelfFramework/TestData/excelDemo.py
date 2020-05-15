import openpyxl
from Tools.scripts.fixcid import Dict

book = openpyxl.load_workbook("C:\\Users\\swetha.priyanka.IDSNEXT\\Desktop\\PythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)

print(sheet.max_column)
print(sheet.max_row)

sheet.cell(row=2, column=2).value = "Rahul"
print(sheet.cell(row=2, column=2).value)
print(sheet['A3'].value)
print("----------------------------------------")
for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "testcase2":
        for j in range(2, sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            #print(sheet.cell(row=i, column=j).value)

print(Dict)